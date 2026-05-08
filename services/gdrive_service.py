from __future__ import annotations

from dataclasses import dataclass
import io
import os
from datetime import datetime, timedelta, timezone
from typing import Any

import httpx
import openpyxl
from docx import Document
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def _drive(access_token: str):
    access_token = (access_token or "").strip()
    if not access_token:
        raise RuntimeError("Missing Google access_token")
    creds = Credentials(
        token=access_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=os.getenv("GOOGLE_CLIENT_ID", ""),
        client_secret=os.getenv("GOOGLE_CLIENT_SECRET", ""),
    )
    return build("drive", "v3", credentials=creds, cache_discovery=True)


def _oauth_client_id() -> str:
    v = (os.getenv("GOOGLE_CLIENT_ID") or "").strip()
    if not v:
        raise RuntimeError("Missing GOOGLE_CLIENT_ID")
    return v


def _oauth_client_secret() -> str:
    v = (os.getenv("GOOGLE_CLIENT_SECRET") or "").strip()
    if not v:
        raise RuntimeError("Missing GOOGLE_CLIENT_SECRET")
    return v


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


@dataclass
class RefreshedToken:
    access_token: str
    expires_at: datetime | None


def refresh_access_token(refresh_token: str) -> RefreshedToken:
    """
    refresh_token으로 access_token 갱신.
    실패 시 예외를 raise합니다.
    """
    refresh_token = (refresh_token or "").strip()
    if not refresh_token:
        raise RuntimeError("Missing refresh_token (re-login required)")

    data = {
        "client_id": _oauth_client_id(),
        "client_secret": _oauth_client_secret(),
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
    }

    r = httpx.post("https://oauth2.googleapis.com/token", data=data, timeout=15.0)
    if r.status_code >= 400:
        raise RuntimeError(f"token refresh failed: {r.status_code} {r.text}")
    j = r.json()
    at = (j.get("access_token") or "").strip()
    if not at:
        raise RuntimeError("token refresh failed: missing access_token")
    expires_in = j.get("expires_in")
    try:
        exp = _utcnow() + timedelta(seconds=int(expires_in or 0))
    except Exception:
        exp = None
    return RefreshedToken(access_token=at, expires_at=exp)


def get_folder_name(folder_id: str, access_token: str) -> str | None:
    folder_id = (folder_id or "").strip()
    if not folder_id:
        return None
    svc = _drive(access_token)
    f = (
        svc.files()
        .get(fileId=folder_id, fields="id,name,mimeType", supportsAllDrives=True)
        .execute()
    )
    if (f.get("mimeType") or "") != "application/vnd.google-apps.folder":
        return None
    return f.get("name")


def list_images_in_folder(folder_id: str, access_token: str, limit: int = 50) -> list[dict[str, Any]]:
    folder_id = (folder_id or "").strip()
    if not folder_id:
        return []
    limit = min(max(int(limit or 50), 1), 200)

    svc = _drive(access_token)
    q = f"'{folder_id}' in parents and trashed = false and mimeType contains 'image/'"
    resp = (
        svc.files()
        .list(
            q=q,
            pageSize=limit,
            fields="files(id,name,mimeType,createdTime,modifiedTime,size,webViewLink,webContentLink)",
            orderBy="modifiedTime desc",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )
    files = resp.get("files", []) or []
    out: list[dict[str, Any]] = []
    for f in files:
        out.append(
            {
                "id": f.get("id"),
                "name": f.get("name"),
                "mimeType": f.get("mimeType"),
                "createdTime": f.get("createdTime"),
                "modifiedTime": f.get("modifiedTime"),
                "size": f.get("size"),
                "webViewLink": f.get("webViewLink"),
                "webContentLink": f.get("webContentLink"),
            }
        )
    return out


def list_videos_in_folder(folder_id: str, access_token: str, limit: int = 10) -> list[dict[str, Any]]:
    folder_id = (folder_id or "").strip()
    if not folder_id:
        return []
    limit = min(max(int(limit or 10), 1), 200)

    svc = _drive(access_token)
    q = f"'{folder_id}' in parents and trashed = false and mimeType contains 'video/'"
    resp = (
        svc.files()
        .list(
            q=q,
            pageSize=limit,
            fields="files(id,name,mimeType,createdTime,modifiedTime,size,webViewLink,webContentLink)",
            orderBy="modifiedTime desc",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )
    files = resp.get("files", []) or []
    out: list[dict[str, Any]] = []
    for f in files:
        out.append(
            {
                "id": f.get("id"),
                "name": f.get("name"),
                "mimeType": f.get("mimeType"),
                "createdTime": f.get("createdTime"),
                "modifiedTime": f.get("modifiedTime"),
                "size": f.get("size"),
                "webViewLink": f.get("webViewLink"),
                "webContentLink": f.get("webContentLink"),
            }
        )
    return out


def get_file_thumbnail_link(file_id: str, access_token: str) -> str | None:
    """Drive files.get thumbnailLink (짧은 캐시 URL; 서버에서만 fetch 권장)."""
    file_id = (file_id or "").strip()
    if not file_id:
        return None
    svc = _drive(access_token)
    meta = (
        svc.files()
        .get(fileId=file_id, fields="thumbnailLink", supportsAllDrives=True)
        .execute()
    )
    return (meta.get("thumbnailLink") or "").strip() or None


def get_parent_folder_id(file_id: str, access_token: str) -> str | None:
    """Direct parent folder id (first parent) for a file or folder."""
    file_id = (file_id or "").strip()
    if not file_id:
        return None
    svc = _drive(access_token)
    meta = (
        svc.files()
        .get(fileId=file_id, fields="parents", supportsAllDrives=True)
        .execute()
    )
    parents = meta.get("parents") or []
    return (parents[0] or "").strip() or None


def download_image(file_id: str, access_token: str) -> tuple[bytes, str]:
    file_id = (file_id or "").strip()
    if not file_id:
        raise ValueError("file_id is required")

    svc = _drive(access_token)
    meta = (
        svc.files()
        .get(fileId=file_id, fields="id,name,mimeType", supportsAllDrives=True)
        .execute()
    )
    mime_type = (meta.get("mimeType") or "application/octet-stream").strip()

    req = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue(), mime_type


def _get_mime_type(file_id: str, access_token: str) -> str:
    svc = _drive(access_token)
    meta = (
        svc.files()
        .get(fileId=file_id, fields="mimeType", supportsAllDrives=True)
        .execute()
    )
    return (meta.get("mimeType") or "").strip()


def read_uploaded_xlsx(file_id: str, access_token: str) -> str | None:
    """xlsx 파일 다운로드 → openpyxl로 텍스트 추출."""
    try:
        data, _mt = download_image(file_id, access_token)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        lines: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                line = "\t".join(str(c or "") for c in row).strip()
                if line:
                    lines.append(line)
        return "\n".join(lines) if lines else None
    except Exception as e:
        print(f"[gdrive] xlsx parse error file_id={file_id}: {e}", flush=True)
        return None


def read_uploaded_docx(file_id: str, access_token: str) -> str | None:
    """docx 파일 다운로드 → python-docx로 텍스트 추출."""
    try:
        data, _mt = download_image(file_id, access_token)
        doc = Document(io.BytesIO(data))
        lines = [p.text.strip() for p in doc.paragraphs if (p.text or "").strip()]
        return "\n".join(lines) if lines else None
    except Exception as e:
        print(f"[gdrive] docx parse error file_id={file_id}: {e}", flush=True)
        return None


_MAX_EXPORT_CHARS = 10000


def _export_workspace_file(
    file_id: str,
    access_token: str,
    expected_mime: str,
    export_mime: str,
) -> str | None:
    file_id = (file_id or "").strip()
    if not file_id:
        return None
    try:
        svc = _drive(access_token)
        meta = (
            svc.files()
            .get(fileId=file_id, fields="id,mimeType", supportsAllDrives=True)
            .execute()
        )
        mime = (meta.get("mimeType") or "").strip()
        if mime != expected_mime:
            print(f"[gdrive] skip file_id={file_id} expected={expected_mime} got={mime}", flush=True)
            return None
        req = svc.files().export_media(fileId=file_id, mimeType=export_mime)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        raw = buf.getvalue().decode("utf-8", errors="replace")
        if not raw:
            return None
        if len(raw) > _MAX_EXPORT_CHARS:
            print(f"[gdrive] truncated file_id={file_id} {len(raw)} → {_MAX_EXPORT_CHARS} chars", flush=True)
            return raw[:_MAX_EXPORT_CHARS]
        return raw
    except Exception as e:
        print(f"[gdrive] export error file_id={file_id} mime={expected_mime}: {e}", flush=True)
        return None


def read_google_doc(file_id: str, access_token: str) -> str | None:
    """Google Docs만 export (text/plain). 일반 파일·비네이티브면 None."""
    return _export_workspace_file(
        file_id,
        access_token,
        "application/vnd.google-apps.document",
        "text/plain",
    )


def read_google_sheet(file_id: str, access_token: str) -> str | None:
    """Google Sheets만 export (text/csv)."""
    return _export_workspace_file(
        file_id,
        access_token,
        "application/vnd.google-apps.spreadsheet",
        "text/csv",
    )


def read_google_slide(file_id: str, access_token: str) -> str | None:
    """Google Slides만 export (text/plain)."""
    return _export_workspace_file(
        file_id,
        access_token,
        "application/vnd.google-apps.presentation",
        "text/plain",
    )


def read_workspace_document(file_id: str, doc_type: str, access_token: str) -> str | None:
    """네이티브/업로드 문서를 mimeType 기반으로 자동 읽기.

    doc_type는 힌트로만 사용하며(legacy), 실제 처리는 Drive mimeType 우선.
    """
    fid = (file_id or "").strip()
    if not fid:
        return None

    try:
        mime = _get_mime_type(fid, access_token)
    except Exception as e:
        print(f"[gdrive] get mimeType failed file_id={fid}: {e}", flush=True)
        return None

    # 네이티브(Google Workspace) → export
    if "application/vnd.google-apps.spreadsheet" in mime:
        return read_google_sheet(fid, access_token)
    if "application/vnd.google-apps.document" in mime:
        return read_google_doc(fid, access_token)
    if "application/vnd.google-apps.presentation" in mime:
        return read_google_slide(fid, access_token)

    # 업로드 파일 → 직접 파싱
    if "spreadsheetml" in mime or "xlsx" in mime:
        return read_uploaded_xlsx(fid, access_token)
    if "wordprocessingml" in mime or "docx" in mime:
        return read_uploaded_docx(fid, access_token)

    # fallback: legacy hint
    dt = (doc_type or "").strip().lower()
    if dt == "sheets":
        return read_google_sheet(fid, access_token)
    if dt == "docs":
        return read_google_doc(fid, access_token)
    if dt == "slides":
        return read_google_slide(fid, access_token)

    print(f"[gdrive] unsupported mime={mime} file_id={fid} doc_type={doc_type}", flush=True)
    return None


